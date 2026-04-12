"""Excel export builder for pricelist positions.

This module ports and cleans up the legacy ``legacy/exceltask.py`` script.

Key behavior:
- Builds an XLSX price list from ``Position`` and ``PriceDate`` tables.
- Stores the file in ``static/excel/pricelist.xlsx``.
- Keeps legacy grouping order: category_id=2, spacer row, category_id=1.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from sqlalchemy import Select, select
from sqlalchemy.orm import Session

from app.modules.pricelist.infrastructure.sa_models import Position, PriceDate
from app.settings.config import settings

OUTPUT_FILENAME = "pricelist.xlsx"

_THIN_BORDER = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)

_HEADER_FONT = Font(name="Calibri", size=12, bold=True)
_BODY_FONT = Font(name="Calibri", size=11)


def _build_group_query(category_id: int) -> Select[tuple[Position]]:
    return (
        select(Position)
        .where(Position.category_id == category_id)
        .order_by(Position.order.asc(), Position.id.asc())
    )


def _string(value: str | None) -> str:
    return (value or "").strip()


def _output_path() -> Path:
    static_root = Path(settings.static.STATIC_ROOT)
    output_dir = static_root / "excel"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / OUTPUT_FILENAME


def generate_pricelist_xlsx(session: Session) -> Path:
    """Generate and save the pricelist XLSX file, returning the output path."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Прайс-лист"

    sheet.column_dimensions["A"].width = 4
    sheet.column_dimensions["B"].width = 55
    sheet.column_dimensions["C"].width = 27
    sheet.column_dimensions["D"].width = 27

    price_date = session.execute(
        select(PriceDate).order_by(PriceDate.id.asc()).limit(1)
    ).scalar_one_or_none()

    sheet["B2"] = "Прайс-лист"
    sheet["B2"].font = Font(name="Calibri", size=14, bold=True)

    sheet["B3"] = f"Дата актуальности: {_string(price_date.date) if price_date else ''}"
    sheet["B3"].font = _BODY_FONT

    headers = [
        "Наименование",
        "Безналичный расчет (на карту физ.лица)",
        "Безналичный расчет (лицензия юр.лица)",
    ]
    header_row = 5
    for col, value in enumerate(headers, start=2):
        cell = sheet.cell(row=header_row, column=col, value=value)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    row = header_row + 1

    for position in session.execute(_build_group_query(category_id=2)).scalars().all():
        sheet.cell(row=row, column=2, value=_string(position.name))
        sheet.cell(row=row, column=3, value=_string(position.price))
        sheet.cell(row=row, column=4, value=_string(position.price_card))
        row += 1

    row += 1  # visual spacer between legacy groups

    for position in session.execute(_build_group_query(category_id=1)).scalars().all():
        sheet.cell(row=row, column=2, value=_string(position.name))
        sheet.cell(row=row, column=3, value=_string(position.price))
        sheet.cell(row=row, column=4, value=_string(position.price_card))
        row += 1

    for data_row in range(header_row + 1, row):
        for col in (2, 3, 4):
            cell = sheet.cell(row=data_row, column=col)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    output_path = _output_path()
    workbook.save(output_path)
    workbook.close()
    return output_path