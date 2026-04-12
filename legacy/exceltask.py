# -*- coding: utf-8 -*-
import os
import sys
from datetime import date

sys.path.insert(0, '/usr/src/vekolom')
sys.path.append("/usr/src/vekolom")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "vekolom.settings")
import django

django.setup()

from pricelist.models import Position, PriceDate
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image

start = date.today()
print("Start time:", start)

gr1 = Position.objects.filter(category_id=2).order_by('order')
gr2 = Position.objects.filter(category_id=1).order_by('order')


wb = load_workbook('/usr/src/vekolom/media/media/exel/PriceVekoStarter.xlsx')
ws = wb.active
try:
    os.remove('/usr/src/vekolom/media/media/exel/PriceVeko.xlsx')
except:
    pass
fill1 = PatternFill(fill_type='solid', start_color='FFFFFFFF', end_color='FF000000')

border1 = Border(left=Side(border_style='thin', color='FF000000'),
                 right=Side(border_style='thin', color='FF000000'),
                 top=Side(border_style='thin', color='FF000000'),
                 bottom=Side(border_style='thin', color='FF000000'))

border2 = Border(left=Side(border_style='thick', color='FF000000'),
                 right=Side(border_style='thick', color='FF000000'),
                 top=Side(border_style='thick', color='FF000000'),
                 bottom=Side(border_style='thick', color='FF000000'))

border3 = Border(top=Side(border_style='thin', color='FF000000'), left=Side(border_style='thin', color='FF000000'))
border4 = Border(left=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
border5 = Border(left=Side(border_style='thin', color='FF000000'))

font1 = Font(name='Calibri', size=14)
font2 = Font(name='Calibri', size=14, italic=True)

price = PriceDate.objects.all()[0]

date = ws['B7']
date.value = price.date
date.font = font1

count1 = 0
for s in gr1:
    count1 += 1
counter1 = 11
for i in gr1:
    refB = 'B' + str(counter1)
    refC = 'C' + str(counter1)
    refD = 'D' + str(counter1)
    k = ws[refB]
    k.value = i.name
    k.border = border1
    k.font = font1
    l = ws[refC]
    l.value = i.price
    l.border = border1
    l.font = font1
    m = ws[refD]
    m.value = i.price_card
    m.border = border1
    m.font = font1
    counter1 += 1

refadd1B = 'B' + str(counter1)
refadd1C = 'C' + str(counter1)
refadd1D = 'D' + str(counter1)
add1B = ws[refadd1B]
add1B.value = ''
add1B.border = border1
add1B.fill = fill1
add1C = ws[refadd1C]
add1C.value = ''
add1C.border = border1
add1C.fill = fill1
add1D = ws[refadd1D]
add1D.value = ''
add1D.border = border1
add1D.fill = fill1

counter2 = counter1 + 1
for i in gr2:
    refB = 'B' + str(counter2)
    refC = 'C' + str(counter2)
    refD = 'D' + str(counter2)
    k = ws[refB]
    k.value = i.name
    k.border = border1
    k.font = font1
    l = ws[refC]
    l.value = i.price
    l.border = border1
    l.font = font1
    d = ws[refD]
    d.value = i.price_card
    d.border = border1
    d.font = font1
    counter2 += 1

img = Image('/usr/src/vekolom/media/media/exel/logoexel.png')
ws.add_image(img, 'C1')

wb.save('/usr/src/vekolom/media/media/exel/PriceVeko.xlsx')
wb.close()
print('excellent')